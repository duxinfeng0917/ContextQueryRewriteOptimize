"""
Query Rewrite Quality Analysis
Reads the dataset, classifies rewrite quality, and outputs a groundtruth evaluation Excel.
"""

import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── helpers ──────────────────────────────────────────────────────────────────

def has_chinese(s):
    return bool(re.search(r'[\u4e00-\u9fff]', s or ''))

def has_latin(s):
    return bool(re.search(r'[a-zA-Z]', s or ''))

def new_latin_words(orig, rew):
    """Return Latin words added in rewrite that weren't in original."""
    orig_words = set(re.findall(r'[A-Za-z]+', orig or ''))
    rew_words  = set(re.findall(r'[A-Za-z]+', rew  or ''))
    return rew_words - orig_words

# Business-term English words that should NOT be injected into Chinese sentences
EN_BUSINESS_TERMS = {
    'monthly', 'payment', 'down', 'ringgit', 'malaysia', 'loan',
    'interest', 'installment', 'deposit', 'balance',
}

# Hallucination markers — spec/detail phrases that should never appear unless in history
HALLUCINATION_PATTERNS = [
    r'model tahun \d{4}',
    r'tahun \d{4}',
    r'warna \w+',
    r'full spec',
    r'spesifikasi penuh',
    r'penawaran yang tersedia',
    r'car models available',
    r'lokasi \w+',
    r'Kuala Lumpur',
]

def extract_dialogue_history(prompt):
    """Extract dialogue history text from the rewrite prompt."""
    if not prompt:
        return ''
    start = prompt.find('# 对话历史')
    end   = prompt.find('# 用户当前消息')
    if start > 0 and end > 0:
        return prompt[start:end].strip()
    return ''

def extract_car_series(prompt):
    """Extract car series hint from the 补充知识 section of the prompt."""
    if not prompt:
        return ''
    # Only look in the 补充知识 section
    section_start = prompt.find('# 补充知识')
    section_end   = prompt.find('# 对话历史')
    if section_start < 0:
        return ''
    section = prompt[section_start:section_end] if section_end > section_start else prompt[section_start:]
    m = re.search(r'对话历史中可能涉及的车系[：:]\s*(.+)', section)
    if not m:
        return ''
    val = m.group(1).strip()
    # Sanity-check: should look like car names (alphanumeric / Chinese), not a section header
    if val.startswith('#') or len(val) > 80:
        return ''
    return val

# ── error detection ───────────────────────────────────────────────────────────

def detect_errors(row):
    """
    Returns a list of (error_code, description) tuples.
    Error codes:
      LANG_ERROR      - language/script changed or business-term injected into Chinese
      OVER_REWRITE    - hallucinated details not present in history
      UNDER_REWRITE   - unresolved pronoun/reference when context is available
      STYLE_REWRITE   - minor wording change without adding info (acceptable but flagged)
      MEANING_CHANGE  - rewrite changes the semantic meaning of the query
    """
    tid, user, intent, uq, trans, lang, sid, rid, uid, orig, translate, ner, prompt, rew, cost = row
    errors = []
    if not orig or not rew:
        return errors

    history = extract_dialogue_history(prompt)

    # ── LANG_ERROR: English business terms injected into Chinese query ──
    if lang == 'Zh' and has_chinese(orig):
        added = new_latin_words(orig, rew)
        bad_terms = {w for w in added if w.lower() in EN_BUSINESS_TERMS}
        if bad_terms:
            errors.append(('LANG_ERROR',
                f'中文query中注入了英文业务词汇: {bad_terms}'))

    # ── LANG_ERROR: Chinese injected into pure Latin query ──
    if lang in ('En', 'My') and not has_chinese(orig) and has_chinese(rew):
        errors.append(('LANG_ERROR', '纯英/马来语query中注入了中文'))

    # ── OVER_REWRITE: hallucinated details ──
    if rew and rew != orig:
        for pat in HALLUCINATION_PATTERNS:
            if re.search(pat, rew, re.IGNORECASE) and not re.search(pat, history, re.IGNORECASE):
                errors.append(('OVER_REWRITE',
                    f'改写添加了对话历史中不存在的细节: "{pat}"'))
                break

    # ── OVER_REWRITE: length explosion on short queries ──
    if rew and orig and len(orig) <= 25 and len(rew) > len(orig) * 2.5:
        # Allow if history is long (more context = longer rewrite might be ok)
        if len(history) < 200:
            errors.append(('OVER_REWRITE',
                f'短query改写过长: orig={len(orig)}chars → rew={len(rew)}chars'))

    # ── MEANING_CHANGE: 首付/月供混淆 ──
    if orig and rew:
        # 首付 (down payment) ≠ 月供 (monthly installment)
        if '首付' in orig and '月供' not in orig:
            if 'Down Payment' in rew and ('Monthly' in rew or '月供' not in rew):
                if '月供' in orig or '月供' not in rew:
                    pass  # not triggered
            # Specific pattern: 第一个月供 changed to Down Payment
            if re.search(r'第.{0,3}月供', orig) and 'Down Payment' in rew:
                errors.append(('MEANING_CHANGE',
                    '"月供"被错误替换为"Down Payment"，语义不同'))

    # ── UNDER_REWRITE: pronoun not resolved despite available context ──
    if orig == rew and history:
        # Only flag truly referential pronouns that appear WITHOUT a co-present car name
        car_hint = extract_car_series(prompt)
        car_in_orig = bool(re.search(r'[A-Z][a-zA-Z0-9]', orig))  # car name likely present

        # Build a set of car names that appear in history/hint
        history_has_car = bool(re.search(r'[A-Z][a-z]+ [A-Z0-9]', history) or car_hint)

        has_unresolved = False

        if lang == 'Zh' and not car_in_orig and history_has_car:
            # Strong referential pronouns in Chinese (subject position or with 辆/款/个)
            strong_zh = ['那辆', '那款', '那个车', '这辆', '这款', '后者', '前者',
                         '它 ', '这辆车', '那辆车']
            if any(p in orig for p in strong_zh):
                has_unresolved = True

        elif lang == 'En' and not car_in_orig and history_has_car:
            # Only flag when pronoun is clearly referential, not filler
            strong_en = [r'\bthe one\b', r'\bthis one\b', r'\bthat one\b',
                         r'\bthe car\b', r'\bthat car\b']
            if any(re.search(p, orig, re.IGNORECASE) for p in strong_en):
                has_unresolved = True

        elif lang in ('My', 'Mix') and not car_in_orig and history_has_car:
            # Strong referential pronouns in Malay (avoid discourse particle "ni/tu")
            # Only flag "kereta ni/tu", "yang ni/tu", "yang Performance tu", not "Awak ni"
            strong_my = [
                r'\bkereta (ni|tu|ini|itu)\b',
                r'\byang (ni|tu|ini|itu)\b',
                r'\byang \w+ tu\b',
                r'\bdia\b',  # "dia" is clearly referential
            ]
            if any(re.search(p, orig, re.IGNORECASE) for p in strong_my):
                has_unresolved = True
            # Mix language: also catch "yg ... tu"
            if lang == 'Mix' and re.search(r'\byg\b.{0,20}\btu\b', orig, re.IGNORECASE):
                has_unresolved = True

        if has_unresolved:
            errors.append(('UNDER_REWRITE',
                '含指代词但未做消解，对话历史中存在可推断的指代对象'))

    return errors


# ── suggest groundtruth ────────────────────────────────────────────────────────

def suggest_groundtruth(row, errors):
    """
    Returns a suggested groundtruth string for error cases.
    For correct cases returns the rewrite result.
    """
    tid, user, intent, uq, trans, lang, sid, rid, uid, orig, translate, ner, prompt, rew, cost = row
    error_codes = [e[0] for e in errors]

    if not errors:
        return rew  # already correct

    history = extract_dialogue_history(prompt)

    # LANG_ERROR in Chinese: strip injected English business terms
    if 'LANG_ERROR' in error_codes and lang == 'Zh':
        # Detect meaning-change: Down Payment used in place of 月供
        # Pattern: "月的 Down Payment" or "月供" replaced by "Down Payment" in rewrite
        if re.search(r'月.{0,5}Down Payment', rew, re.IGNORECASE) and '月供' not in rew:
            return f'[需人工] "Down Payment"错误替换了"月供"，语义不同，原句: {orig}'
        if re.search(r'第.{0,3}月.{0,5}Down Payment', rew, re.IGNORECASE):
            return f'[需人工] "Down Payment"错误替换了"月供"，语义不同，原句: {orig}'

        suggestion = rew
        replacements = [
            (r'Monthly Payment', '月供'),
            (r'Monthly Installment', '月供'),
            (r'Monthly', '月供'),
            (r'Down Payment', '首付'),
            (r'\s*\(mthly\)', ''),
            (r'\bmthly\b', ''),
            (r'20,000 Ringgit Malaysia', '2万'),
            (r'Ringgit Malaysia', '马币'),
            (r'20,000', '2万'),
            (r'\s+how much\??', '多少？'),
        ]
        for pat, rep in replacements:
            suggestion = re.sub(pat, rep, suggestion, flags=re.IGNORECASE)
        # Deduplicate adjacent identical Chinese words (e.g. "月供 月供")
        suggestion = re.sub(r'([\u4e00-\u9fff]{2,4}) \1', r'\1', suggestion)
        suggestion = re.sub(r'\s+', ' ', suggestion).strip()
        # Verify no more injected business terms
        added = new_latin_words(orig, suggestion)
        if {w for w in added if w.lower() in EN_BUSINESS_TERMS}:
            return f'[需人工] 建议保留中文措辞，参考原句: {orig}'
        return suggestion

    # MEANING_CHANGE: 月供/首付混淆
    if 'MEANING_CHANGE' in error_codes:
        return f'[需人工] 语义错误，原句: {orig}'

    # OVER_REWRITE: strip hallucinated additions
    if 'OVER_REWRITE' in error_codes:
        # For cases where original is already self-contained, keep original
        if re.search(r'[A-Z][a-z]+ \w+', orig):  # has car name
            return orig
        # Otherwise use original (the hallucinated details shouldn't be added)
        return f'[需人工] 过度改写，建议参考对话历史仅补全指代，原句: {orig}'

    # UNDER_REWRITE: suggest resolving pronoun
    if 'UNDER_REWRITE' in error_codes:
        car_hint = extract_car_series(prompt)
        if car_hint:
            return f'[需人工] 建议补全指代 ({car_hint})，原句: {orig}'
        return f'[需人工] 建议从对话历史补全指代，原句: {orig}'

    return f'[需人工] {orig}'


# ── classification ────────────────────────────────────────────────────────────

def classify(row):
    """Return (label, error_summary, suggested_gt)."""
    tid, user, intent, uq, trans, lang, sid, rid, uid, orig, translate, ner, prompt, rew, cost = row

    errors = detect_errors(row)
    error_codes = [e[0] for e in errors]

    if orig == rew:
        if errors:
            label = 'UNDER_REWRITE'
        else:
            label = 'NO_CHANGE'  # standalone query, no rewrite needed
    else:
        if errors:
            # Most severe error takes priority
            priority = ['MEANING_CHANGE', 'LANG_ERROR', 'OVER_REWRITE', 'UNDER_REWRITE', 'STYLE_REWRITE']
            label = next((p for p in priority if p in error_codes), error_codes[0])
        else:
            label = 'CORRECT'

    error_summary = '; '.join(f'{e[0]}: {e[1]}' for e in errors) if errors else ''
    suggested_gt = suggest_groundtruth(row, errors)
    return label, error_summary, suggested_gt


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    wb_in = openpyxl.load_workbook('dataset/马来西亚新车数据集.xlsx')
    ws_in = wb_in['Sheet']
    all_rows = list(ws_in.iter_rows(min_row=2, values_only=True))

    # ── Build output workbook ──
    wb_out = openpyxl.Workbook()

    # Sheet 1: Full evaluation table
    ws_eval = wb_out.active
    ws_eval.title = '改写质量评测'

    headers = [
        '测试ID', '语言', 'intent', '原始query', '改写结果',
        '质量标签', '错误说明', '建议groundtruth',
        '人工审核结论', '人工修正groundtruth', '备注'
    ]
    ws_eval.append(headers)

    # Color map
    fill_map = {
        'CORRECT':       PatternFill('solid', fgColor='C6EFCE'),
        'NO_CHANGE':     PatternFill('solid', fgColor='DDEBF7'),
        'UNDER_REWRITE': PatternFill('solid', fgColor='FFEB9C'),
        'LANG_ERROR':    PatternFill('solid', fgColor='FFC7CE'),
        'OVER_REWRITE':  PatternFill('solid', fgColor='FFCC99'),
        'MEANING_CHANGE':PatternFill('solid', fgColor='FF6B6B'),
        'STYLE_REWRITE': PatternFill('solid', fgColor='F2F2F2'),
    }

    stats = {}
    for row in all_rows:
        label, error_summary, suggested_gt = classify(row)
        stats[label] = stats.get(label, 0) + 1

        tid = row[0]; lang = row[5]; intent = row[2]
        orig = row[9]; rew = row[13]

        out_row = [
            tid, lang, intent, orig, rew,
            label, error_summary, suggested_gt,
            '', '', ''  # human review columns
        ]
        ws_eval.append(out_row)
        # Apply color to label cell (column F = 6)
        last_row = ws_eval.max_row
        fill = fill_map.get(label, PatternFill('solid', fgColor='FFFFFF'))
        ws_eval.cell(row=last_row, column=6).fill = fill

    # Style header
    header_fill = PatternFill('solid', fgColor='4472C4')
    bold_white = Font(bold=True, color='FFFFFF')
    for col in range(1, len(headers) + 1):
        cell = ws_eval.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = bold_white
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Column widths
    widths = [14, 6, 22, 40, 40, 16, 60, 60, 16, 60, 30]
    for i, w in enumerate(widths, 1):
        ws_eval.column_dimensions[get_column_letter(i)].width = w

    ws_eval.freeze_panes = 'A2'
    ws_eval.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

    # Sheet 2: Statistics
    ws_stats = wb_out.create_sheet('统计总览')
    ws_stats.append(['质量标签', '数量', '占比%', '说明'])
    label_desc = {
        'CORRECT':        '改写正确：正确补全了指代/缩写，未过度改写',
        'NO_CHANGE':      '无需改写：原句独立完整，保持原样输出',
        'UNDER_REWRITE':  '漏改写：含指代词但未消解，历史中有可推断对象',
        'LANG_ERROR':     '语种错误：改写引入了与原句语种不一致的词汇',
        'OVER_REWRITE':   '过度改写：添加了对话历史中不存在的细节',
        'MEANING_CHANGE': '语义错误：改写改变了原句的语义意图',
        'STYLE_REWRITE':  '样式改写：仅做了措辞润色，未增加实质信息',
    }
    total = len(all_rows)
    for lbl, cnt in sorted(stats.items(), key=lambda x: -x[1]):
        pct = round(cnt / total * 100, 1)
        ws_stats.append([lbl, cnt, pct, label_desc.get(lbl, '')])

    ws_stats.append([])
    ws_stats.append(['总计', total, 100.0, ''])

    # Style stats sheet
    for col in range(1, 5):
        ws_stats.cell(row=1, column=col).fill = header_fill
        ws_stats.cell(row=1, column=col).font = bold_white
    ws_stats.column_dimensions['A'].width = 18
    ws_stats.column_dimensions['B'].width = 8
    ws_stats.column_dimensions['C'].width = 8
    ws_stats.column_dimensions['D'].width = 55

    # Sheet 3: Error-only subset (for quick review)
    ws_errors = wb_out.create_sheet('错误案例')
    error_labels = {'LANG_ERROR', 'OVER_REWRITE', 'MEANING_CHANGE', 'UNDER_REWRITE'}
    ws_errors.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws_errors.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = bold_white
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Re-run classification to populate error sheet
    for row in all_rows:
        label, error_summary, suggested_gt = classify(row)
        if label in error_labels:
            tid = row[0]; lang = row[5]; intent = row[2]
            orig = row[9]; rew = row[13]
            out_row = [tid, lang, intent, orig, rew, label, error_summary, suggested_gt, '', '', '']
            ws_errors.append(out_row)
            last_row = ws_errors.max_row
            fill = fill_map.get(label, PatternFill('solid', fgColor='FFFFFF'))
            ws_errors.cell(row=last_row, column=6).fill = fill

    for i, w in enumerate(widths, 1):
        ws_errors.column_dimensions[get_column_letter(i)].width = w
    ws_errors.freeze_panes = 'A2'

    out_path = 'dataset/groundtruth_eval.xlsx'
    wb_out.save(out_path)
    print(f'Saved: {out_path}')
    print('\n=== Quality Statistics ===')
    for lbl, cnt in sorted(stats.items(), key=lambda x: -x[1]):
        pct = cnt / total * 100
        print(f'  {lbl:<18} {cnt:>4}  ({pct:.1f}%)')
    print(f'  {"Total":<18} {total:>4}')

    # Print error cases for inspection
    error_count = sum(v for k, v in stats.items() if k in error_labels)
    print(f'\nError cases requiring fix: {error_count} ({error_count/total*100:.1f}%)')


if __name__ == '__main__':
    main()
