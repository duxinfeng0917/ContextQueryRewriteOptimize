"""
dspy_optimizer.py
DSPy 自动提示词优化脚本

功能：
  1. 从 groundtruth_eval.xlsx + 原始数据集 加载有标注的改写样本
  2. 用程序化 metric 评估改写质量（语言一致、无幻觉、指代消解）
  3. 通过 DSPy MIPROv2 自动优化指令 + 生成 few-shot 示例
  4. 将优化后程序保存为 JSON，可随时加载复用

模型：
  gpt-5.4-nano，通过 llm_infer/litellm_client.py 统一管理调用

用法：
  python3 dspy_optimizer.py                # 完整 MIPROv2 优化
  python3 dspy_optimizer.py --quick        # 快速模式（BootstrapFewShot）
  python3 dspy_optimizer.py --eval         # 仅评估，不优化
  python3 dspy_optimizer.py --load optimized_rewriter.json --eval

依赖：
  pip install dspy-ai openpyxl litellm openai
"""

import re
import os
import sys
import json
import random
import argparse
import openpyxl
import dspy
from dspy.evaluate import Evaluate

# LiteLLM 统一客户端（模型调用 / DSPy LM 构造）
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'llm_infer'))
from litellm_client import make_dspy_lm  # noqa: E402

_DEFAULT_MODEL = 'gpt-5.4-nano'

# ── 固定缩写词表（不参与优化，作为输入字段传入）──────────────────────────────
ABBREVIATION_TABLE = """\
| 缩写 | 完整写法 |
|------|----------|
| mthly | Monthly |
| dp | Down Payment |
| fl | Full Loan |
| rm | Ringgit Malaysia |
| k | Thousand（如 50k = 50,000）|
| nego | Negotiable |
| cond | Condition |
| yr | Year |
| mil | Mileage |
| acc free | Accident Free |
| ori | Original |
| mod | Modified |
| fs | Full Spec |
| ss | Standard Spec |
| recon | Reconditioned |
| mfg | Manufacturing Year |
| reg | Registered Year |
| SA | Sales Advisor |
| doc | Document |
| ic | Identity Card |
| lesen | Lesen Memandu |
| stmt | Bank Statement |
| epf | Employees Provident Fund |
| bl | Blacklist |
| guar | Guarantor |
| int | Interest Rate |
| HP | Hire Purchase |
| depo | Deposit |
| bal | Balance |
| ins | Insurance |
| jpj | Jabatan Pengangkutan Jalan |
| loc | Location |
| tq | Thank You |
| ws | WhatsApp |
| wasap | WhatsApp |
| brp | berapa |
| hrga | harga |
| kete | kereta |
| thn | tahun |
| bln | bulan |
| sy | saya |
| nk | nak |
| x | tak |
| bleh | boleh |
| skrg | sekarang |
| dh | sudah |
| dah | sudah |
| blm | belum |
| utk | untuk |
| jg | juga |
| tu | itu |
| ni | ini |
| kt | dekat |
| tau | tahu |
| cntk | cantik |
| msh | masih |
| tgu | tunggu |
| pki | pakai |
| kwsp | Kumpulan Wang Simpanan Pekerja |"""

# ── 错误检测工具（复用 analyze_rewrites.py 的逻辑）────────────────────────────

EN_BUSINESS_TERMS = {
    'monthly', 'payment', 'down', 'ringgit', 'malaysia', 'loan',
    'interest', 'installment', 'deposit', 'balance',
}

HALLUCINATION_PATTERNS = [
    r'model tahun \d{4}',
    r'tahun \d{4}',
    r'warna \w+',
    r'full spec',
    r'spesifikasi penuh',
    r'penawaran yang tersedia',
    r'car models available',
    r'lokasi \w+',
    r'Kuala Lumpur',
]


def has_chinese(s: str) -> bool:
    return bool(re.search(r'[一-鿿]', s or ''))


def has_latin(s: str) -> bool:
    return bool(re.search(r'[a-zA-Z]', s or ''))


def new_latin_words(orig: str, rew: str) -> set:
    orig_words = set(re.findall(r'[A-Za-z]+', orig or ''))
    rew_words = set(re.findall(r'[A-Za-z]+', rew or ''))
    return rew_words - orig_words


def check_lang_error(orig: str, predicted: str, lang: str) -> bool:
    """返回 True 表示存在语种错误。"""
    if lang == 'Zh' and has_chinese(orig):
        added = new_latin_words(orig, predicted)
        if {w for w in added if w.lower() in EN_BUSINESS_TERMS}:
            return True
    if lang in ('En', 'My') and not has_chinese(orig) and has_chinese(predicted):
        return True
    return False


def check_over_rewrite(orig: str, predicted: str, history: str) -> bool:
    """返回 True 表示过度改写（幻觉）。"""
    if predicted == orig:
        return False
    for pat in HALLUCINATION_PATTERNS:
        if re.search(pat, predicted, re.IGNORECASE) and not re.search(pat, history, re.IGNORECASE):
            return True
    # 短 query 改写过长
    if len(orig) <= 25 and len(predicted) > len(orig) * 2.5 and len(history) < 200:
        return True
    return False


def check_under_rewrite(orig: str, predicted: str, lang: str, history: str, car_series: str) -> bool:
    """返回 True 表示漏改写（指代未消解）。"""
    if orig != predicted:
        return False  # 已经做了改写，不算漏
    if not history:
        return False

    car_in_orig = bool(re.search(r'[A-Z][a-zA-Z0-9]', orig))
    history_has_car = bool(re.search(r'[A-Z][a-z]+ [A-Z0-9]', history) or car_series)
    if not history_has_car:
        return False

    if lang == 'Zh' and not car_in_orig:
        strong_zh = ['那辆', '那款', '那个车', '这辆', '这款', '后者', '前者',
                     '这辆车', '那辆车']
        if any(p in orig for p in strong_zh):
            return True

    elif lang == 'En' and not car_in_orig:
        strong_en = [r'\bthe one\b', r'\bthis one\b', r'\bthat one\b',
                     r'\bthe car\b', r'\bthat car\b']
        if any(re.search(p, orig, re.IGNORECASE) for p in strong_en):
            return True

    elif lang in ('My', 'Mix') and not car_in_orig:
        strong_my = [
            r'\bkereta (ni|tu|ini|itu)\b',
            r'\byang (ni|tu|ini|itu)\b',
            r'\byang \w+ tu\b',
            r'\bdia\b',
        ]
        if any(re.search(p, orig, re.IGNORECASE) for p in strong_my):
            return True
        if lang == 'Mix' and re.search(r'\byg\b.{0,20}\btu\b', orig, re.IGNORECASE):
            return True

    return False


# ── Prompt 解析工具 ──────────────────────────────────────────────────────────

def extract_history(prompt: str) -> str:
    """从 rewrite_prompt 中提取对话历史文本。"""
    if not prompt:
        return ''
    start = prompt.find('# 对话历史')
    end = prompt.find('# 用户当前消息')
    if start >= 0 and end > start:
        raw = prompt[start + len('# 对话历史'):end].strip()
        return raw
    return ''


def extract_car_series_from_prompt(prompt: str) -> str:
    """从 rewrite_prompt 的补充知识段落提取车系信息。"""
    if not prompt:
        return ''
    section_start = prompt.find('# 补充知识')
    section_end = prompt.find('# 对话历史')
    if section_start < 0:
        return ''
    section = prompt[section_start:section_end] if section_end > section_start else prompt[section_start:]
    m = re.search(r'对话历史中可能涉及的车系[：:]\s*(.+)', section)
    if not m:
        return ''
    val = m.group(1).strip()
    if val.startswith('#') or len(val) > 80:
        return ''
    return val


def extract_current_query(prompt: str) -> str:
    """从 rewrite_prompt 中提取用户当前消息。"""
    if not prompt:
        return ''
    start = prompt.find('# 用户当前消息')
    if start < 0:
        return ''
    section = prompt[start + len('# 用户当前消息'):]
    m = re.search(r'User:\s*(.+?)(?:\n|$)', section)
    if m:
        return m.group(1).strip()
    return ''


# ── DSPy Signature 与 Module ──────────────────────────────────────────────────

class QueryRewriteSignature(dspy.Signature):
    """
    多语言 query 改写：结合对话历史，补全用户问题中的指代信息，展开常见缩写词。

    硬性规则：
    1. 【禁止改变语种】输出必须与 current_query 使用相同语言/脚本（中→中，英→英，马来→马来）
    2. 【只用已有信息】只能使用 dialogue_history 或 car_series 中明确出现的信息来补全
    3. 【只做补全+展开】不添加解释、建议、推理过程；不修改句式或纠正语法
    4. 【信息不足则保留】若历史中无足够信息确定指代对象，保持 current_query 原样输出
    5. 【缩写展开】按 abbreviation_table 展开；车型名/专有名词中的字母（如 X50、S70）不展开
    """

    car_series: str = dspy.InputField(desc="对话中可能涉及的车系（可为空字符串）")
    dialogue_history: str = dspy.InputField(desc="完整对话历史，包含 User/Assistant 轮次")
    language_type: str = dspy.InputField(desc="用户消息的语言类型：Zh / En / My / Mix")
    abbreviation_table: str = dspy.InputField(desc="缩写词表（固定不变，按此表展开缩写）")
    current_query: str = dspy.InputField(desc="用户当前待改写的消息")

    rewritten_query: str = dspy.OutputField(
        desc="改写结果：在保持原语言的前提下，补全指代信息并展开缩写词。"
             "若无需改写则输出原句。只输出改写后的问题本身，不要任何前缀或解释。"
    )


class QueryRewriter(dspy.Module):
    def __init__(self):
        super().__init__()
        self.predict = dspy.Predict(QueryRewriteSignature)

    def forward(
        self,
        car_series: str,
        dialogue_history: str,
        language_type: str,
        current_query: str,
    ) -> dspy.Prediction:
        return self.predict(
            car_series=car_series,
            dialogue_history=dialogue_history,
            language_type=language_type,
            abbreviation_table=ABBREVIATION_TABLE,
            current_query=current_query,
        )


# ── 评估 Metric ───────────────────────────────────────────────────────────────

def rewrite_metric(example: dspy.Example, pred: dspy.Prediction, trace=None) -> float:
    """
    改写质量评分 [0, 1]。
    评分策略：
      - 语言错误 → 0.0（硬失败）
      - 过度改写（幻觉）→ 0.0（硬失败）
      - 与 groundtruth 精确匹配 → 1.0
      - NO_CHANGE 案例且输出等于原句 → 1.0
      - CORRECT/NO_CHANGE 无误但不精确匹配 → 0.6
      - 错误案例（LANG/OVER/UNDER）且无检测到错误 → 0.8
      - 漏改写（UNDER_REWRITE）但实际做了合理改写 → 0.7
    """
    predicted = (pred.rewritten_query or '').strip()
    if not predicted:
        return 0.0

    orig = example.current_query
    lang = example.language_type
    history = example.dialogue_history
    car_series = example.car_series
    label = example.expected_label
    expected_gt = (example.expected_gt or '').strip()

    # ── 硬失败检查 ──────────────────────────────────────────────────
    if check_lang_error(orig, predicted, lang):
        return 0.0

    if check_over_rewrite(orig, predicted, history):
        return 0.0

    # ── 精确匹配 groundtruth ─────────────────────────────────────────
    if expected_gt and not expected_gt.startswith('[需人工]'):
        if predicted == expected_gt:
            return 1.0

    # ── 按标签分类评分 ───────────────────────────────────────────────
    if label == 'NO_CHANGE':
        # 原句独立完整，不应改写
        return 1.0 if predicted == orig else 0.4

    elif label == 'CORRECT':
        # 已正确改写；只要无错误即可得 0.6，精确匹配才得 1.0
        has_under = check_under_rewrite(orig, predicted, lang, history, car_series)
        if has_under:
            return 0.3
        return 0.6

    elif label == 'UNDER_REWRITE':
        # 期望做了指代消解
        if predicted != orig:
            # 做了改写且无新错误
            return 0.8
        return 0.2

    elif label in ('LANG_ERROR', 'OVER_REWRITE', 'MEANING_CHANGE'):
        # 已通过硬失败检查，说明预测没有这些错误
        return 0.8

    return 0.5


# ── 数据加载 ──────────────────────────────────────────────────────────────────

def load_examples(
    eval_path: str = 'dataset/groundtruth_eval.xlsx',
    orig_path: str = 'dataset/马来西亚新车数据集.xlsx',
    include_needs_manual: bool = False,
) -> list[dspy.Example]:
    """
    加载并合并评测集与原始数据集，构建 DSPy Example 列表。

    参数：
      include_needs_manual: 是否包含 suggested_gt 为 "[需人工]" 的案例
                            （这类案例没有自动 groundtruth，仅用于评估，不用于训练）
    """
    # 加载原始数据集（获取 rewrite_prompt 列）
    orig_wb = openpyxl.load_workbook(orig_path)
    orig_ws = orig_wb['Sheet']
    prompt_map: dict[str, str] = {}
    for row in orig_ws.iter_rows(min_row=2, values_only=True):
        tid = row[0]
        rewrite_prompt = row[12]
        if tid and rewrite_prompt:
            prompt_map[tid] = str(rewrite_prompt)

    # 加载评测结果
    eval_wb = openpyxl.load_workbook(eval_path)
    eval_ws = eval_wb['改写质量评测']

    examples = []
    for row in eval_ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        tid = row[0]
        lang = row[1] or 'Zh'
        orig_query = str(row[3] or '').strip()
        label = str(row[5] or 'CORRECT').strip()
        suggested_gt = str(row[7] or '').strip()
        manual_gt = str(row[9] or '').strip()

        if not tid or not orig_query:
            continue

        # groundtruth 优先使用人工修正版
        gt = manual_gt if manual_gt else suggested_gt

        # 跳过没有可用 groundtruth 的案例（除非显式要求包含）
        if not include_needs_manual and gt.startswith('[需人工]'):
            continue

        # 从原始 prompt 中提取结构化字段
        prompt = prompt_map.get(tid, '')
        history = extract_history(prompt)
        car_series = extract_car_series_from_prompt(prompt)
        current_query = extract_current_query(prompt) or orig_query

        examples.append(
            dspy.Example(
                car_series=car_series,
                dialogue_history=history,
                language_type=lang,
                current_query=current_query,
                expected_gt=gt,
                expected_label=label,
            ).with_inputs('car_series', 'dialogue_history', 'language_type', 'current_query')
        )

    return examples


def split_data(
    examples: list[dspy.Example],
    train_ratio: float = 0.7,
    val_ratio: float = 0.15,
    seed: int = 42,
) -> tuple[list, list, list]:
    """按 CORRECT/NO_CHANGE/错误 分层采样，保证各类别分布均衡。"""
    random.seed(seed)

    # 分层
    correct = [e for e in examples if e.expected_label == 'CORRECT']
    no_change = [e for e in examples if e.expected_label == 'NO_CHANGE']
    errors = [e for e in examples if e.expected_label not in ('CORRECT', 'NO_CHANGE')]

    def split_group(group, train_r, val_r):
        random.shuffle(group)
        n = len(group)
        n_train = int(n * train_r)
        n_val = int(n * val_r)
        return group[:n_train], group[n_train:n_train + n_val], group[n_train + n_val:]

    c_tr, c_val, c_te = split_group(correct, train_ratio, val_ratio)
    n_tr, n_val, n_te = split_group(no_change, train_ratio, val_ratio)
    e_tr, e_val, e_te = split_group(errors, train_ratio, val_ratio)

    train = c_tr + n_tr + e_tr
    val = c_val + n_val + e_val
    test = c_te + n_te + e_te

    random.shuffle(train)
    random.shuffle(val)
    random.shuffle(test)

    return train, val, test


# ── 主流程 ────────────────────────────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(description='DSPy query rewrite optimizer')
    parser.add_argument('--quick', action='store_true',
                        help='快速模式：使用 BootstrapFewShot（省 token，适合测试）')
    parser.add_argument('--eval', action='store_true',
                        help='仅评估，不执行优化')
    parser.add_argument('--load', type=str, default='',
                        help='加载已保存的优化程序路径，不重新优化')
    parser.add_argument('--max-train', type=int, default=200,
                        help='训练集最大样本数（控制优化成本，默认 200）')
    parser.add_argument('--seed', type=int, default=42)
    parser.add_argument('--output', type=str, default='optimized_rewriter.json',
                        help='优化结果保存路径')
    return parser.parse_args()


def _make_lm(max_tokens: int = 512) -> dspy.LM:
    """构造 DSPy LM 实例（通过 LiteLLM 管理调用）。"""
    return make_dspy_lm(_DEFAULT_MODEL, max_tokens=max_tokens)


def main():
    args = parse_args()
    random.seed(args.seed)

    # ── 配置 DSPy LM（gpt-5.4-nano via Azure）──────────────────────
    lm = _make_lm(max_tokens=2048)
    dspy.configure(lm=lm)
    print(f'[LM] 模型: {_DEFAULT_MODEL}（via LiteLLM）')

    # ── 加载数据 ──────────────────────────────────────────────────────
    print('\n[Data] 加载评测集...')
    examples = load_examples()
    print(f'[Data] 可用样本: {len(examples)}  (排除了 [需人工] 案例)')

    label_counts = {}
    for e in examples:
        label_counts[e.expected_label] = label_counts.get(e.expected_label, 0) + 1
    print('[Data] 标签分布:')
    for k, v in sorted(label_counts.items(), key=lambda x: -x[1]):
        print(f'       {k:<18} {v:>4}')

    train_set, val_set, test_set = split_data(examples, seed=args.seed)

    # 限制训练集大小（控制优化成本）
    if len(train_set) > args.max_train:
        # 保证各类别都有代表
        correct_tr = [e for e in train_set if e.expected_label == 'CORRECT']
        nc_tr = [e for e in train_set if e.expected_label == 'NO_CHANGE']
        err_tr = [e for e in train_set if e.expected_label not in ('CORRECT', 'NO_CHANGE')]
        budget_per_class = args.max_train // 3
        train_set = (
            correct_tr[:budget_per_class]
            + nc_tr[:budget_per_class]
            + err_tr[:min(len(err_tr), args.max_train - 2 * budget_per_class)]
        )
        random.shuffle(train_set)

    print(f'\n[Data] 训练集: {len(train_set)}  验证集: {len(val_set)}  测试集: {len(test_set)}')

    # ── 初始化程序 ────────────────────────────────────────────────────
    program = QueryRewriter()

    if args.load:
        print(f'\n[Load] 加载已保存的程序: {args.load}')
        program.load(args.load)

    # ── 基线评估 ──────────────────────────────────────────────────────
    if not args.load:
        print('\n[Eval] 基线评估（未优化）...')
        evaluator = Evaluate(
            devset=val_set[:50],
            metric=rewrite_metric,
            num_threads=4,
            display_progress=True,
        )
        baseline_score = float(evaluator(program))
        print(f'[Eval] 基线验证集得分: {baseline_score:.3f}')

    # ── 优化 ──────────────────────────────────────────────────────────
    if not args.eval and not args.load:
        if args.quick:
            print('\n[Opt] 快速模式：BootstrapFewShot (省 token，适合快速验证)')
            optimizer = dspy.BootstrapFewShot(
                metric=rewrite_metric,
                max_bootstrapped_demos=4,
                max_labeled_demos=4,
                max_rounds=1,
            )
            optimized = optimizer.compile(program, trainset=train_set)

        else:
            print(f'\n[Opt] MIPROv2 优化（{_DEFAULT_MODEL}）...')
            optimizer = dspy.MIPROv2(
                metric=rewrite_metric,
                auto='medium',           # 自动平衡质量与成本
                num_threads=4,
                verbose=True,
            )
            optimized = optimizer.compile(
                program,
                trainset=train_set,
                valset=val_set,
                max_bootstrapped_demos=3,
                max_labeled_demos=4,
                requires_permission_to_run=False,
            )

        # 保存优化结果
        optimized.save(args.output)
        print(f'\n[Save] 优化程序已保存至: {args.output}')

        # 优化后评估
        print('\n[Eval] 优化后评估...')
        evaluator = Evaluate(
            devset=val_set,
            metric=rewrite_metric,
            num_threads=4,
            display_progress=True,
        )
        opt_score = float(evaluator(optimized))
        print(f'[Eval] 优化后验证集得分: {opt_score:.3f}')

        # 测试集最终评估
        print('\n[Eval] 测试集最终评估...')
        test_evaluator = Evaluate(
            devset=test_set,
            metric=rewrite_metric,
            num_threads=4,
            display_progress=True,
        )
        test_score = float(test_evaluator(optimized))
        print(f'[Eval] 测试集得分: {test_score:.3f}')

        program = optimized

    elif args.load:
        # 仅评估加载的程序
        print('\n[Eval] 评估加载的程序...')
        evaluator = Evaluate(
            devset=test_set,
            metric=rewrite_metric,
            num_threads=4,
            display_progress=True,
        )
        score = float(evaluator(program))
        print(f'[Eval] 测试集得分: {score:.3f}')

    # ── 交互式推理演示 ─────────────────────────────────────────────────
    print('\n' + '='*60)
    print('[Demo] 随机选取 3 个样本展示改写效果')
    print('='*60)
    demo_samples = random.sample(test_set[:50] if len(test_set) >= 50 else test_set, min(3, len(test_set)))
    for i, ex in enumerate(demo_samples, 1):
        pred = program(
            car_series=ex.car_series,
            dialogue_history=ex.dialogue_history,
            language_type=ex.language_type,
            current_query=ex.current_query,
        )
        score = rewrite_metric(ex, pred)
        print(f'\n--- 样本 {i} [{ex.language_type}] label={ex.expected_label} ---')
        if ex.dialogue_history:
            hist_preview = ex.dialogue_history[:200].replace('\n', ' ')
            print(f'  历史: {hist_preview}...' if len(ex.dialogue_history) > 200 else f'  历史: {hist_preview}')
        print(f'  车系: {ex.car_series}')
        print(f'  原句: {ex.current_query}')
        print(f'  预测: {pred.rewritten_query}')
        print(f'  期望: {ex.expected_gt}')
        print(f'  得分: {score:.2f}')

    print('\n[Done] 优化流程完成。')
    if not args.eval and not args.load:
        print(f'       已保存优化程序至: {args.output}')
        print(f'       加载方式: program.load("{args.output}")')


if __name__ == '__main__':
    main()
